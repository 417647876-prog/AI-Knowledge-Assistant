from pathlib import Path

from openpyxl import load_workbook

from app.core.exceptions import AppError
from app.knowledge.schemas import ParsedSection


class ExcelParser:
    def parse(self, file_path: Path) -> list[ParsedSection]:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sections: list[ParsedSection] = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(value or "") for value in rows[0]]
            for index, row in enumerate(rows[1:], start=2):
                text = "\n".join(
                    f"{header}: {value}"
                    for header, value in zip(headers, row, strict=False)
                    if value is not None
                )
                if text:
                    sections.append(
                        ParsedSection(text=text, sheet_name=sheet.title, row_start=index)
                    )
        if not sections:
            raise AppError(code="DOCUMENT_CONTENT_EMPTY", message="文档内容为空。", status_code=422)
        return sections
